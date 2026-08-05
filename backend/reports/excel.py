import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def generate_excel_file(
    data,
    title="HRMS Report",
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet["A1"] = title
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    if not data:
        worksheet["A3"] = "No data available"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output.getvalue()

    headers = list(data[0].keys())

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=3,
            column=column_index,
            value=str(header).replace("_", " ").title(),
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
        )

    for row_index, row in enumerate(
        data,
        start=4,
    ):
        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            value = row.get(header)

            if hasattr(value, "isoformat"):
                value = value.isoformat()

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        max_length = len(str(header))

        for cell in worksheet[
            get_column_letter(column_index)
        ]:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = min(max_length + 3, 40)

    worksheet.freeze_panes = "A4"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()